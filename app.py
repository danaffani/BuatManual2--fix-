import time
from flask import Flask, render_template_string, request, redirect, jsonify, send_from_directory, url_for
import webview
import openpyxl
import json
import threading
import os
import requests
from datetime import datetime

import signal   
import sys

def signal_handler(signal, frame):
    print('Exiting...')
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)

app = Flask(__name__)
app.secret_key = 'your_secret_key'

file_path = 'database/item.xlsx'
person_file_path = 'database/person.xlsx'
sales_file_path = 'database/sales.xlsx'

def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    headers = [cell.value for cell in sheet[1]]  # Mengambil header dari baris pertama
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: str(row[i]) if headers[i].endswith('_id') else row[i] for i in range(len(row))}
        data.append(item)
    return data

def write_excel(data, file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = list(data[0].keys())  # Convert dict_keys to a list
    sheet.append(headers)  # Create headers in the first row

    for item in data:
        sheet.append(list(item.values()))

    workbook.save(file_path)

def get_next_item_id(data, key):
    if not data:
        return 1
    return max(int(item[key]) for item in data) + 1

def run_flask():
    app.run(debug=True, use_reloader=False)

@app.route('/')
def home():
    return render_template_string(r'''
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <nav class="navbar navbar-expand-lg navbar-light bg-light">
            <a class="navbar-brand" href="/">Market App</a>
            <div class="collapse navbar-collapse">
                <ul class="navbar-nav mr-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="/stock">Stock</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/person">Person</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/sales">Sale</a>
                    </li>
                </ul>
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item">
                        <a class="btn btn-danger btn-sm delete-button" href="/exit" onclick="return confirm('Are you sure you want to exit?');">Exit</a>
                    </li>
                </ul>
            </div>
        </nav>
        <div class="container mt-5">
            <h1 class="text-center mt-4">Welcome to the Market App</h1>
        </div>
        <script src="https://code.jquery.com/jquery-3.5.1.slim.min.js"></script>
        <script src="https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js"></script>
        <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
        <script>
            window.addEventListener('beforeunload', function (event) {
                navigator.sendBeacon('/shutdown');
            });
        </script>
    ''')

@app.route('/welcome', methods=['GET', 'POST'])
def welcome():
    return render_template_string(r'''
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <nav class="navbar navbar-expand-lg navbar-light bg-light">
            <a class="navbar-brand" href="/">Market App</a>
            <div class="collapse navbar-collapse">
                <ul class="navbar-nav mr-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="/stock">Stock</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/person">Person</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/sales">Sale</a>
                    </li>
                </ul>
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item">
                        <a class="btn btn-danger btn-sm delete-button" href="/exit" onclick="return confirm('Are you sure you want to exit?');">Exit</a>
                    </li>
                </ul>
            </div>
        </nav>
        <div class="container mt-5">
            <h1 class="text-center">Welcome to Market App</h1>
        </div>
        <script src="https://code.jquery.com/jquery-3.5.1.slim.min.js"></script>
        <script src="https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js"></script>
        <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
        <script>
            window.addEventListener('beforeunload', function (event) {
                navigator.sendBeacon('/shutdown');
            });
        </script>
    ''')

@app.route('/stock', methods=['GET', 'POST'])
def stock():
    if request.method == 'POST':
        if request.form.get('action') == 'delete':
            item_id = request.form['item_id']
            data = [item for item in read_excel(file_path) if item['item_id'] != item_id]
            write_excel(data, file_path)
            return redirect('/stock')

    data = read_excel(file_path)
    return render_template_string(r'''
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <nav class="navbar navbar-expand-lg navbar-light bg-light">
            <a class="navbar-brand" href="/">Market App</a>
            <div class="collapse navbar-collapse">
                <ul class="navbar-nav mr-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="/stock">Stock</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/person">Person</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/sales">Sale</a>
                    </li>
                </ul>
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item">
                        <a class="btn btn-danger btn-sm delete-button" href="/exit" onclick="return confirm('Are you sure you want to exit?');">Exit</a>
                    </li>
                </ul>
            </div>
        </nav>
        <div class="container mt-5">
            <h1 class="text-center">Stock Management</h1>
            <div class="text-right mb-3">
                <a href="/open-stock-folder" class="btn btn-secondary">Open Folder</a>
                <a href="/edit" class="btn btn-primary">Add New Data</a>
            </div>
            <div class="form-group mb-3">
                <input type="text" id="stock-search-box" class="form-control" placeholder="Search items..." onkeyup="searchStockTable()">
            </div>
            <table class="table table-bordered table-striped">
                <thead class="thead-light">
                    <tr>
                        <th>Item ID</th>
                        <th>Item Name</th>
                        <th>Item Price</th>
                        <th>Item Available</th>
                        <th>Item Sold</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody id="stock-table-body">
                    {% for item in data %}
                    <tr>
                        <td>{{ item.item_id }}</td>
                        <td>{{ item.item_name }}</td>
                        <td class="item-price">{{ item.item_price }}</td>
                        <td>{{ item.item_available }}</td>
                        <td>{{ item.item_sold }}</td>
                        <td>
                            <a href="/edit?id={{ item.item_id }}" class="btn btn-warning btn-sm">Edit</a>
                            <form method="POST" style="display:inline;" class="delete-form">
                                <input type="hidden" name="item_id" value="{{ item.item_id }}">
                                <button type="submit" name="action" value="delete" class="btn btn-danger btn-sm delete-button">Delete</button>
                            </form>
                        </td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        <script src="https://code.jquery.com/jquery-3.5.1.min.js"></script>
        <script src="https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js"></script>
        <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
        <script>        
            function formatCurrency(value) {
                let number = Math.floor(parseFloat(value));
                return 'Rp. ' + number.toString().replace(/\B(?=(\d{3})+(?!\d))/g, ".");
            }
            
            function searchStockTable() {
                let input = document.getElementById('stock-search-box').value.toLowerCase();
                let rows = document.querySelectorAll('#stock-table-body tr');

                rows.forEach(row => {
                    let cells = row.querySelectorAll('td');
                    let rowText = Array.from(cells).map(cell => cell.textContent.toLowerCase()).join(' ');
                    row.style.display = rowText.includes(input) ? '' : 'none';
                });
            }

            document.addEventListener('DOMContentLoaded', function() {
                let priceElements = document.querySelectorAll('.item-price');
                priceElements.forEach(function(element) {
                    let price = parseFloat(element.innerText.replace(/[^0-9.]/g, ''));
                    element.innerText = formatCurrency(price.toFixed(2));
                });

                const deleteForms = document.querySelectorAll('.delete-form');
                deleteForms.forEach(form => {
                    form.addEventListener('submit', function(event) {
                        if (!confirm('Are you sure you want to delete this item?')) {
                            event.preventDefault();
                        }
                    });
                });
            });
                                  
            window.onbeforeunload = function () {
                fetch('/shutdown', {method: 'POST'});
            };
            
            window.addEventListener('beforeunload', function (event) {
    navigator.sendBeacon('/shutdown');
});
        </script>
    ''', data=data)

@app.route('/open-stock-folder')
def open_stock_folder():
    folder_path = os.path.dirname(os.path.abspath(file_path))
    os.startfile(folder_path)
    return redirect('/stock')

@app.route('/edit', methods=['GET', 'POST'])
def edit():
    data = read_excel(file_path)
    if request.method == 'POST':
        item_id = request.form['item_id']
        if item_id:
            for item in data:
                if item['item_id'] == item_id:
                    item['item_name'] = request.form['item_name']
                    item['item_price'] = request.form['item_price']
                    item['item_available'] = request.form['item_available']
                    break
        else:
            new_item = {
                'item_id': str(get_next_item_id(data, 'item_id')),
                'item_name': request.form['item_name'],
                'item_price': request.form['item_price'],
                'item_available': request.form['item_available'],
                'item_sold': '0'
            }
            data.append(new_item)
        write_excel(data, file_path)
        return redirect('/stock')

    item_id = request.args.get('id')
    item = next((item for item in data if item['item_id'] == item_id), None)
    return render_template_string(r'''
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <nav class="navbar navbar-expand-lg navbar-light bg-light">
            <a class="navbar-brand" href="/">Market App</a>
            <div class="collapse navbar-collapse">
                <ul class="navbar-nav mr-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="/stock">Stock</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/person">Person</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/sales">Sale</a>
                    </li>
                </ul>
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item">
                        <a class="btn btn-danger btn-sm delete-button" href="/exit" onclick="return confirm('Are you sure you want to exit?');">Exit</a>
                    </li>
                </ul>
            </div>
        </nav>
        <div class="container mt-5">
            <h1 class="text-center">Edit Item</h1>
            <form method="POST">
                <div class="form-group">
                    <label for="item_name">Item Name:</label>
                    <input type="text" class="form-control" id="item_name" name="item_name" value="{{ item.item_name if item else '' }}" required>
                </div>
                <div class="form-group">
                    <label for="item_price">Item Price:</label>
                    <input type="number" class="form-control" id="item_price" name="item_price" value="{{ item.item_price if item else '' }}" required>
                </div>
                <div class="form-group">
                    <label for="item_available">Item Available:</label>
                    <input type="number" class="form-control" id="item_available" name="item_available" value="{{ item.item_available if item else '' }}" required>
                </div>
                <input type="hidden" name="item_id" value="{{ item.item_id if item else '' }}">
                <button type="submit" class="btn btn-primary">Save</button>
                <a href="/stock" class="btn btn-secondary">Cancel</a>
            </form>
        </div>
        <script src="https://code.jquery.com/jquery-3.5.1.min.js"></script>
        <script src="https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js"></script>
        <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
        <script>            
            window.addEventListener('beforeunload', function (event) {
                navigator.sendBeacon('/shutdown');
            });
        </script>
    ''', item=item)

@app.route('/person', methods=['GET', 'POST'])
def person():
    data = read_excel(person_file_path)
    if request.method == 'POST' and request.form.get('action') == 'delete':
            person_id = request.form['person_id']
            data = [person for person in data if person['person_id'] != person_id]
            write_excel(data, person_file_path)
            return redirect('/person')
    return render_template_string(r'''
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <nav class="navbar navbar-expand-lg navbar-light bg-light">
            <a class="navbar-brand" href="/">Market App</a>
            <div class="collapse navbar-collapse">
                <ul class="navbar-nav mr-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="/stock">Stock</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/person">Person</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/sales">Sale</a>
                    </li>
                </ul>
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item">
                        <a class="btn btn-danger btn-sm delete-button" href="/exit" onclick="return confirm('Are you sure you want to exit?');">Exit</a>
                    </li>
                </ul>
            </div>
        </nav>
        <div class="container mt-5">
            <h1 class="text-center">Person Management</h1>
            <div class="text-right mb-3">
                <a href="/open-person-folder" class="btn btn-secondary">Open Folder</a>
                <a href="/edit_person" class="btn btn-primary">Add New Person</a>
            </div>
            <div class="form-group mb-3">
                <input type="text" id="person-search-box" class="form-control" placeholder="Search persons..." onkeyup="searchPersonTable()">
            </div>
            <table class="table table-bordered table-striped">
                <thead class="thead-light">
                    <tr>
                        <th>Person ID</th>
                        <th>Person Name</th>
                        <th>Visit Time</th>
                        <th>Purchased Item</th>
                        <th>Money Spent</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody id="person-table-body">
                    {% for person in data %}
                    <tr>
                        <td>{{ person.person_id }}</td>
                        <td>{{ person.person_name }}</td>
                        <td>{{ person.visit_time }}</td>
                        <td>{{ person.purchased_item }}</td>
                        <td class="money-spent">{{ person.money_spent }}</td>
                        <td>
                            <a href="/edit_person?id={{ person.person_id }}" class="btn btn-warning btn-sm">Edit</a>
                            <form method="POST" style="display:inline;" class="delete-form" onsubmit="return confirmDelete()">
                                <input type="hidden" name="person_id" value="{{ person.person_id }}">
                                <input type="hidden" name="action" value="delete">
                                <button type="submit" name="action" value="delete" class="btn btn-danger btn-sm delete-button">Delete</button>
                            </form>
                            <a href="/details?id={{ person.person_id }}" class="btn btn-info btn-sm">Details</a>
                        </td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        <script src="https://code.jquery.com/jquery-3.5.1.min.js"></script>
        <script src="https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js"></script>
        <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
        <script>
            function searchPersonTable() {
                let input = document.getElementById('person-search-box').value.toLowerCase();
                let rows = document.querySelectorAll('#person-table-body tr');
                rows.forEach(row => {
                    let cells = row.querySelectorAll('td');
                    let rowText = Array.from(cells).map(cell => cell.textContent.toLowerCase()).join(' ');
                    row.style.display = rowText.includes(input) ? '' : 'none';
                });
            }
            
            function formatCurrency(value) {
                let number = Math.floor(parseFloat(value));
                return 'Rp. ' + number.toString().replace(/\B(?=(\d{3})+(?!\d))/g, ".");
            }
            
            function confirmDelete() {
                return confirm('Are you sure you want to delete this person?');
            }

            document.addEventListener('DOMContentLoaded', function() {
                let moneySpentElements = document.querySelectorAll('.money-spent');
                moneySpentElements.forEach(function(element) {
                    let moneySpent = parseFloat(element.innerText.replace(/[^0-9.]/g, ''));
                    element.innerText = formatCurrency(moneySpent);
                });

                let itemTotalPriceElements = document.querySelectorAll('.item-total-price');
                itemTotalPriceElements.forEach(function(element) {
                    let itemTotalPrice = parseFloat(element.innerText.replace(/[^0-9.]/g, ''));
                    element.innerText = formatCurrency(itemTotalPrice);
                });
            });
            
            window.addEventListener('beforeunload', function (event) {
    navigator.sendBeacon('/shutdown');
});
        </script>
    ''', data=data)
    
@app.route('/open-person-folder')
def open_person_folder():
    folder_path = os.path.dirname(os.path.abspath(person_file_path))
    os.startfile(folder_path)
    return redirect('/person')

@app.route('/edit_person', methods=['GET', 'POST'])
def edit_person():
    person_id = request.args.get('id')
    person_data = read_excel(person_file_path)
    person = next((p for p in person_data if p['person_id'] == person_id), None)

    if request.method == 'POST':
        new_id = request.form['person_id']
        new_name = request.form['person_name']
        existing_person = next((p for p in person_data if p['person_id'] == new_id), None)

        if existing_person and new_id != person_id:
            return render_template_string('''
                <p>ID Person sudah ada, milik {{ existing_person.person_name }}. Silakan masukkan ID lain.</p>
                <a href="/edit_person?id={{ person_id }}">Kembali ke Edit Person</a>
            ''', existing_person=existing_person, person_id=person_id)

        # Update person data
        if person:
            person['person_id'] = new_id
            person['person_name'] = new_name
        else:
            new_person = {
                'person_id': new_id,
                'person_name': new_name,
                'visit_time': 0,  # Initialize visit_time to 0
                'purchased_item': 0,  # Initialize purchased_item to 0
                'money_spent': 0.0  # Initialize money_spent to 0.0
            }
            person_data.append(new_person)

        # Update sales data if person_id changed
        if person_id != new_id:
            sales_data = read_excel(sales_file_path)
            for sale in sales_data:
                if sale['person_id'] == person_id:
                    sale['person_id'] = new_id
            write_excel(sales_data, sales_file_path)

        write_excel(person_data, person_file_path)
        return redirect('/person')

    data = read_excel(person_file_path)
    person_id = request.args.get('id')
    person = next((person for person in data if person['person_id'] == person_id), None)
    return render_template_string(r'''
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <nav class="navbar navbar-expand-lg navbar-light bg-light">
            <a class="navbar-brand" href="/">Market App</a>
            <div class="collapse navbar-collapse">
                <ul class="navbar-nav mr-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="/stock">Stock</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/person">Person</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/sales">Sale</a>
                    </li>
                </ul>
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item">
                        <a class="btn btn-danger btn-sm delete-button" href="/exit" onclick="return confirm('Are you sure you want to exit?');">Exit</a>
                    </li>
                </ul>
            </div>
        </nav>
        <div class="container mt-5">
            <h1 class="text-center">Edit Person</h1>
            <form method="POST">
                <div class="form-group">
                    <label for="person_id">Person ID:</label>
                    <input type="text" class="form-control" id="person_id" name="person_id" value="{{ person.person_id if person else '' }}" required>
                </div>
                <div class="form-group">
                    <label for="person_name">Person Name:</label>
                    <input type="text" class="form-control" id="person_name" name="person_name" value="{{ person.person_name if person else '' }}" required>
                </div>
                <input type="hidden" name="person_id" value="{{ person.person_id if person else '' }}">
                <button type="submit" class="btn btn-primary">Save</button>
                <a href="/person" class="btn btn-secondary">Cancel</a>
            </form>
        </div>
        <script src="https://code.jquery.com/jquery-3.5.1.min.js"></script>
        <script src="https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js"></script>
        <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
        <script>
            window.addEventListener('beforeunload', function (event) {
                navigator.sendBeacon('/shutdown');
            });
        </script>
    ''', person=person, person_id=person_id)

@app.route('/details', methods=['GET'])
def details():
    person_id = request.args.get('id')
    person_data = read_excel(person_file_path)
    sales_data = read_excel(sales_file_path)
    stock_data = read_excel(file_path)

    person = next((p for p in person_data if p['person_id'] == person_id), None)
    person_sales = [sale for sale in sales_data if sale['person_id'] == person_id]

    for sale in person_sales:
        item = next((i for i in stock_data if i['item_id'] == sale['item_id']), None)
        if item:
            sale['item_name'] = item['item_name']
            
    person_sales = sorted(person_sales, key=lambda x: x['purchase_date'], reverse=True)

    return render_template_string(r'''
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <nav class="navbar navbar-expand-lg navbar-light bg-light">
            <a class="navbar-brand" href="/">Market App</a>
            <div class="collapse navbar-collapse">
                <ul class="navbar-nav mr-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="/stock">Stock</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/person">Person</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/sales">Sale</a>
                    </li>
                </ul>
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item">
                        <a class="btn btn-danger btn-sm delete-button" href="/exit" onclick="return confirm('Are you sure you want to exit?');">Exit</a>
                    </li>
                </ul>
            </div>
        </nav>
        <div class="container mt-5">
            <h1 class="text-center">Purchase History</h1>
            <div class="mb-3">
                <a href="/person" class="btn btn-secondary">Back</a>
            </div>
            <div class="form-group mb-3">
                <input type="text" id="sales-search-box" class="form-control" placeholder="Search sales..." onkeyup="searchSalesTable()">
            </div>
            <table class="table table-bordered table-striped">
                <thead class="thead-light">
                    <tr>
                        <th>Sale ID</th>
                        <th>Item Name</th>
                        <th>Item Quantity</th>
                        <th>Item Total Price</th>
                        <th>Purchase Date</th>
                    </tr>
                </thead>
                <tbody id="sales-table-body">
                    {% for sale in person_sales %}
                    <tr>
                        <td>{{ sale.sale_id }}</td>
                        <td>{{ sale.item_name }}</td>
                        <td>{{ sale.item_quantity }}</td>
                        <td class="item-total-price">{{ sale.item_total_price }}</td>
                        <td>{{ sale.purchase_date }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        <script src="https://code.jquery.com/jquery-3.5.1.min.js"></script>
        <script src="https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js"></script>
        <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
        <script>
            function searchSalesTable() {
                let input = document.getElementById('sales-search-box').value.toLowerCase();
                let rows = document.querySelectorAll('#sales-table-body tr');
                rows.forEach(row => {
                    let cells = row.querySelectorAll('td');
                    let rowText = Array.from(cells).map(cell => cell.textContent.toLowerCase()).join(' ');
                    row.style.display = rowText.includes(input) ? '' : 'none';
                });
            }
            
            function formatCurrency(value) {
                let number = Math.floor(parseFloat(value));
                return 'Rp. ' + number.toString().replace(/\B(?=(\d{3})+(?!\d))/g, ".");
            }

            document.addEventListener('DOMContentLoaded', function() {
                let moneySpentElements = document.querySelectorAll('.money-spent');
                moneySpentElements.forEach(function(element) {
                    let moneySpent = parseFloat(element.innerText.replace(/[^0-9.]/g, ''));
                    element.innerText = formatCurrency(moneySpent);
                });

                let itemTotalPriceElements = document.querySelectorAll('.item-total-price');
                itemTotalPriceElements.forEach(function(element) {
                    let itemTotalPrice = parseFloat(element.innerText.replace(/[^0-9.]/g, ''));
                    element.innerText = formatCurrency(itemTotalPrice);
                });
            });
            
            window.addEventListener('beforeunload', function (event) {
                navigator.sendBeacon('/shutdown');
            });
        </script>
    ''', person_sales=person_sales)
    
@app.route('/rake_up', methods=['GET', 'POST'])
def rake_up():
    sales_data = read_excel(sales_file_path)
    items_data = read_excel(file_path)  # Baca data dari item.xlsx
    sold_items = {item['item_id']: {'item_name': item['item_name'], 'total_sold': 0} for item in items_data}  # Inisialisasi dictionary untuk menyimpan jumlah terjual

    if request.method == 'POST':
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        
        # Hitung jumlah terjual berdasarkan rentang tanggal
        for sale in sales_data:
            purchase_date = datetime.strptime(sale['purchase_date'], '%d/%m/%Y-%H:%M:%S')
            if start_date <= purchase_date.strftime('%Y-%m-%d') <= end_date:
                item_id = sale['item_id']
                if item_id in sold_items:
                    sold_items[item_id]['total_sold'] += sale['item_quantity']

    # Filter dan siapkan data untuk ditampilkan
    filtered_sold_items = {item_id: data for item_id, data in sold_items.items() if data['total_sold'] > 0}

    return render_template_string(r'''
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <nav class="navbar navbar-expand-lg navbar-light bg-light">
            <a class="navbar-brand" href="/">Market App</a>
            <div class="collapse navbar-collapse">
                <ul class="navbar-nav mr-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="/stock">Stock</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/person">Person</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/rake_up">Rake Up</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/sales">Sale</a>
                    </li>
                </ul>
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item">
                        <a class="btn btn-danger btn-sm delete-button" href="/exit" onclick="return confirm('Are you sure you want to exit?');">Exit</a>
                    </li>
                </ul>
            </div>
        </nav>
        <div class="container mt-5">
            <h1 class="text-center">Sales Rake Up</h1>
            <form method="POST" class="mb-4">
                <div class="form-row">
                    <div class="form-group col-md-6">
                        <label for="start_date">Since:</label>
                        <input type="date" class="form-control" id="start_date" name="start_date" required>
                    </div>
                    <div class="form-group col-md-6">
                        <label for="end_date">Until:</label>
                        <input type="date" class="form-control" id="end_date" name="end_date" required>
                    </div>
                </div>
                <button type="submit" class="btn btn-primary">Tampilkan</button>
            </form>
            <table class="table table-bordered table-striped">
                <thead class="thead-light">
                    <tr>
                        <th>Sale ID</th>
                        <th>Buyer</th>
                        <th>Item Name</th>
                        <th>Item Quantity</th>
                        <th>Item Total Price</th>
                        <th>Purchase Date</th>
                    </tr>
                </thead>
                <tbody>
                    {% for sale in filtered_sales %}
                    <tr>
                        <td>{{ sale.sale_id }}</td>
                        <td>{{ sale.person_name }}</td>
                        <td>{{ sale.item_name }}</td>
                        <td>{{ sale.item_quantity }}</td>
                        <td>{{ sale.item_total_price }}</td>
                        <td>{{ sale.purchase_date }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        <script src="https://code.jquery.com/jquery-3.5.1.min.js"></script>
        <script src="https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js"></script>
        <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
    ''', filtered_sales=filtered_sold_items)

@app.route('/sales', methods=['GET', 'POST'])
def sales():
    if request.method == 'POST':
        data = request.json
        person_id = data.get('person_id')
        cart_items = data.get('cart_items')

        if not cart_items:
            return jsonify({'success': False, 'message': 'No items in cart.'})

        sales_data = read_excel(sales_file_path)
        person_data = read_excel(person_file_path)
        stock_data = read_excel(file_path)

        person = next((p for p in person_data if p['person_id'] == person_id), None)

        if person:
            
            if sales_data:
                last_sale_id = max(int(sale['sale_id']) for sale in sales_data)
            else:
                last_sale_id = 0
            new_sale_id = last_sale_id + 1
            
            total_quantity = 0
            total_price = 0
            for item in cart_items:
                item_id = item['item_id']
                item_quantity = int(item['quantity'])
                item_total_price = float(item['price']) * item_quantity
                purchase_date = datetime.now().strftime('%d/%m/%Y-%H:%M:%S')

                sales_data.append({
                    'sale_id': new_sale_id,
                    'person_id': person_id,
                    'item_id': item_id,
                    'item_quantity': item_quantity,
                    'item_total_price': item_total_price,
                    'purchase_date': purchase_date
                })

                total_quantity += item_quantity
                total_price += item_total_price

                stock_item = next((i for i in stock_data if i['item_id'] == item_id), None)
                if stock_item:
                    stock_item['item_available'] = str(int(stock_item['item_available']) - item_quantity)

            person['visit_time'] = str(int(person['visit_time']) + 1)
            person['purchased_item'] = str(int(person['purchased_item']) + total_quantity)
            person['money_spent'] = str(float(person['money_spent']) + total_price)

            write_excel(sales_data, sales_file_path)
            write_excel(person_data, person_file_path)
            write_excel(stock_data, file_path)

            return jsonify({'success': True})

    data = read_excel(file_path)
    person_data = read_excel(person_file_path)
    return render_template_string(r'''
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <link href="https://cdn.jsdelivr.net/npm/select2@4.1.0-rc.0/dist/css/select2.min.css" rel="stylesheet" />
        <nav class="navbar navbar-expand-lg navbar-light bg-light">
            <a class="navbar-brand" href="/">Market App</a>
            <div class="collapse navbar-collapse">
                <ul class="navbar-nav mr-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="/stock">Stock</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/person">Person</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/sales">Sale</a>
                    </li>
                </ul>
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item">
                        <a class="btn btn-danger btn-sm delete-button" href="/exit" onclick="return confirm('Are you sure you want to exit?');">Exit</a>
                    </li>
                </ul>
            </div>
        </nav>
        <div class="container-fluid mt-5"> <!-- Use container-fluid for full-width layout -->
            <h1 class="text-center">Sales</h1>
            <div class="row">
                <div class="col-md-7">
                    <h2>Stock Item List</h2>
                    <div class="form-group">
                        <input type="text" id="search-box" class="form-control" placeholder="Search list items..." onkeyup="searchTable('items-table')">
                    </div>
                    <div class="table-responsive"> <!-- Add table-responsive for better handling on smaller screens -->
                        <table class="table table-bordered table-striped">
                            <thead class="thead-light">
                                <tr>
                                    <th>Item Name</th>
                                    <th>Item Available</th>
                                    <th>Jumlah</th>
                                    <th>Actions</th>
                                </tr>
                            </thead>
                            <tbody id="items-table">
                                {% for item in data %}
                                <tr>
                                    <td>{{ item.item_name }}</td>
                                    <td>{{ item.item_available }}</td>
                                    <td>
                                        <div class="input-group">
                                            <div class="input-group-prepend">
                                                <button onclick="decreaseQuantity('{{ item.item_id }}')" onmousedown="startDecrease('{{ item.item_id }}')" onmouseup="stopChange()" onmouseleave="stopChange()" class="btn btn-secondary">-</button>
                                            </div>
                                            <input type="number" id="quantity-{{ item.item_id }}" value="0" min="0" max="{{ item.item_available }}" class="form-control" readonly>
                                            <div class="input-group-append">
                                                <button onclick="increaseQuantity('{{ item.item_id }}')" onmousedown="startIncrease('{{ item.item_id }}')" onmouseup="stopChange()" onmouseleave="stopChange()" class="btn btn-secondary">+</button>
                                            </div>
                                        </div>
                                    </td>
                                    <td>
                                        <button onclick="addToCart('{{ item.item_id }}', '{{ item.item_name }}', '{{ item.item_price }}')" class="btn btn-primary">Add to Cart</button>
                                    </td>
                                </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>
                </div>
                <div class="col-md-5">
                    <h2>Cart Item List</h2>
                    <div class="form-group">
                        <input type="text" id="cart-search-box" class="form-control" placeholder="Search cart items..." onkeyup="searchTable('cart-items')">
                    </div>
                    <div class="table-responsive">
                        <table id="cart-table" class="table table-bordered table-striped">
                            <thead class="thead-light">
                                <tr>
                                    <th>Item Name</th>
                                    <th>Jumlah</th>
                                    <th>Item Price</th>
                                    <th>Item Total Price</th>
                                    <th>Actions</th>
                                </tr>
                            </thead>
                            <tbody id="cart-items">
                                <!-- Cart items will be populated here -->
                            </tbody>
                        </table>
                    </div>
                    <h3 id="total-price" style="display: none;">Total Price: Rp. 0</h3>
                    <div class="form-group">
                        <label for="person_id">Select Customer:</label>
                        <select class="form-control" id="person_id" name="person_id" required>
                            <option value="">Select a person...</option>
                            {% for person in person_data %}
                            <option value="{{ person.person_id }}">{{ person.person_name }} ({{ person.person_id }})</option>
                            {% endfor %}
                        </select>
                    </div>
                    <button onclick="jual()" class="btn btn-success mt-3">Sale</button>
                    <button id="refresh-button" class="btn btn-secondary mt-3" style="display: none;" onclick="refreshPage()">Refresh</button>

                    <div class="mt-4" "mb-4">
                        <h2 style="margin-top: 20px;">Receipt:</h2>
                        <div id="receipt" style="display: none; background-color: white; padding: 20px; border: 1px solid #ccc; margin-bottom: 20px;">
                            <h3>MarketApp Receipt</h3>
                            <p id="receipt-timestamp"></p>
                            <p id="receipt-customer"></p>
                            <table class="table">
                                <thead>
                                    <tr>
                                        <th>Item Name</th>
                                        <th>Quantity</th>
                                        <th>Price</th>
                                        <th>Total</th>
                                    </tr>
                                </thead>
                                <tbody id="receipt-items">
                                    <!-- Receipt items will be populated here -->
                                </tbody>
                            </table>
                            <h4 id="receipt-total-price">Total Price: Rp. 0</h4>
                        </div>
                        <button id="download-receipt" class="btn btn-info mt-3" hidden>Download Receipt</button>
                    </div>
                    <a href="/open-receipt-folder" class="btn btn-secondary">Open Receipt Folder</a>
                    <a href="/open-sale-folder" class="btn btn-secondary">Open Sale Folder</a>
                </div>
            </div>
        </div>

        <script src="https://code.jquery.com/jquery-3.5.1.min.js"></script>
        <script src="https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js"></script>
        <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
        <script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/0.4.1/html2canvas.min.js"></script>
        <script src="https://cdn.jsdelivr.net/npm/select2@4.1.0-rc.0/dist/js/select2.min.js"></script>
        <script>
            let interval;
            const cart = {};
            
            document.addEventListener('DOMContentLoaded', function() {
                $('#person_id').select2({
                    placeholder: 'Search for a person...',
                    allowClear: true
                });
            });

            function formatCurrency(value) {
                let number = Math.floor(parseFloat(value));
                return 'Rp. ' + number.toString().replace(/\B(?=(\d{3})+(?!\d))/g, ".");
            }

            function startIncrease(itemId) {
                interval = setInterval(() => {
                    let input = document.getElementById('quantity-' + itemId);
                    input.value = Math.min(parseInt(input.value) + 1, parseInt(input.max));
                }, 100);
            }

            function startDecrease(itemId) {
                interval = setInterval(() => {
                    let input = document.getElementById('quantity-' + itemId);
                    input.value = Math.max(parseInt(input.value) - 1, 0);
                }, 100);
            }

            function stopChange() {
                clearInterval(interval);
            }

            function increaseQuantity(itemId) {
                let input = document.getElementById('quantity-' + itemId);
                input.value = Math.min(parseInt(input.value) + 1, parseInt(input.max));
            }

            function decreaseQuantity(itemId) {
                let input = document.getElementById('quantity-' + itemId);
                input.value = Math.max(parseInt(input.value) - 1, 0);
            }

            function addToCart(itemId, itemName, itemPrice) {
                let quantityInput = document.getElementById('quantity-' + itemId);
                let quantity = parseInt(quantityInput.value);

                if (quantity > 0) {
                    if (cart[itemId]) {
                        cart[itemId].quantity += quantity;
                    } else {
                        cart[itemId] = {
                            item_id: itemId,
                            name: itemName,
                            price: parseFloat(itemPrice),
                            quantity: quantity
                        };
                    }
                    updateCart();
                    quantityInput.value = 0;
                }
            }

            function updateCart() {
                let cartItems = document.getElementById('cart-items');
                cartItems.innerHTML = '';

                let totalPrice = 0;

                for (let itemId in cart) {
                    let item = cart[itemId];
                    let totalItemPrice = (item.price * item.quantity).toFixed(2);
                    totalPrice += parseFloat(totalItemPrice);

                    let row = document.createElement('tr');
                    row.innerHTML = `
                        <td>${item.name}</td>
                        <td>${item.quantity}</td>
                        <td>${formatCurrency(item.price.toFixed(2))}</td>
                        <td>${formatCurrency(totalItemPrice)}</td>
                        <td>
                            <button onclick="removeFromCart('${itemId}')" class="btn btn-danger">Remove</button>
                        </td>
                    `;
                    cartItems.appendChild(row);
                }

                let totalPriceElement = document.getElementById('total-price');
                if (totalPrice > 0) {
                    totalPriceElement.style.display = 'block';
                    totalPriceElement.innerText = 'Total Price: ' + formatCurrency(totalPrice.toFixed(2));
                } else {
                    totalPriceElement.style.display = 'none';
                }
            }

            function removeFromCart(itemId) {
                delete cart[itemId];
                updateCart();
            }

            function displayReceipt(cartItems, personName, personId) {
                const receipt = document.getElementById('receipt');
                const receiptItems = document.getElementById('receipt-items');
                const receiptTimestamp = document.getElementById('receipt-timestamp');
                const receiptCustomer = document.getElementById('receipt-customer');
                const receiptTotalPrice = document.getElementById('receipt-total-price');
                const downloadReceiptButton = document.getElementById('download-receipt');

                receiptItems.innerHTML = '';
                let total = 0;

                for (let itemId in cart) {
                    const item = cart[itemId];
                    const itemTotal = item.price * item.quantity;
                    total += itemTotal;

                    const row = document.createElement('tr');
                    row.innerHTML = `
                        <td>${item.name}</td>
                        <td>${item.quantity}</td>
                        <td>${formatCurrency(item.price.toFixed(2))}</td>
                        <td>${formatCurrency(itemTotal.toFixed(2))}</td>
                    `;
                    receiptItems.appendChild(row);
                }

                const now = new Date();
                const day = String(now.getDate()).padStart(2, '0');
                const month = String(now.getMonth() + 1).padStart(2, '0');
                const year = now.getFullYear();
                const hours = String(now.getHours()).padStart(2, '0');
                const minutes = String(now.getMinutes()).padStart(2, '0');
                const seconds = String(now.getSeconds()).padStart(2, '0');
                const formattedDate = `${day}${month}${year}`;
                const formattedTime = `${hours}${minutes}${seconds}`;
                const formattedDate1 = `${day}/${month}/${year}`;
                const formattedTime1 = `${hours}:${minutes}:${seconds}`;
                const filename = `receipt-${formattedDate}-${formattedTime}.png`;

                receiptTimestamp.innerText = `Purchase Time: ${formattedDate1}, ${formattedTime1}`;
                receiptCustomer.innerText = `Customer: ${personName}`;
                receiptTotalPrice.innerText = `Total Price: ${formatCurrency(total.toFixed(2))}`;

                receipt.style.display = 'block';
                downloadReceiptButton.style.display = 'block';

                html2canvas(receipt, {
                    onrendered: function(canvas) {
                        canvas.toBlob(function(blob) {
                            const formData = new FormData();
                            formData.append('receipt', blob, filename);
                            
                            $.ajax({
                                url: '/save_receipt',
                                type: 'POST',
                                data: formData,
                                processData: false,
                                contentType: false,
                                success: function(response) {
                                    console.log('Receipt saved successfully');
                                },
                                error: function() {
                                    alert('Error saving receipt.');
                                }
                            });
                        });
                    }
                });

                downloadReceiptButton.onclick = function() {
                    html2canvas(receipt).then(function(canvas) {
                        const link = document.createElement('a');
                        link.download = filename;
                        link.href = canvas.toDataURL('image/png');
                        link.click();
                    });
                };
            }

            function jual() {
                const personId = document.getElementById('person_id').value;
                const cartItems = Object.values(cart);
                const personSelect = document.getElementById('person_id');
                const selectedPersonId = personSelect.value;
                const selectedPersonName = personSelect.options[personSelect.selectedIndex].text;

                fetch('/sales', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify({
                        person_id: selectedPersonId,
                        person_name: selectedPersonName,
                        cart_items: cartItems
                    })
                })
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        alert('Sale successful!');
                        displayReceipt(cartItems, selectedPersonName, selectedPersonId);

                        const jualButton = document.querySelector('button.btn-success');
                        const refreshButton = document.getElementById('refresh-button');

                        jualButton.disabled = true;
                        jualButton.classList.add('disabled');
                        refreshButton.style.display = 'block';                        
                    } else {
                        alert('Sale failed!');
                    }
                })
                .catch(error => {
                    console.error('Error:', error);
                });
            }

            function refreshPage() {
                window.location.reload();
            }
                                  
            function searchTable(tableId) {
                let input = document.getElementById(tableId === 'items-table' ? 'search-box' : 'cart-search-box').value.toLowerCase();
                let rows = document.querySelectorAll(`#${tableId} tr`);

                rows.forEach(row => {
                    let cells = row.querySelectorAll('td');
                    let rowText = Array.from(cells).map(cell => cell.textContent.toLowerCase()).join(' ');
                    row.style.display = rowText.includes(input) ? '' : 'none';
                });
            }
            
            window.addEventListener('beforeunload', function (event) {
    navigator.sendBeacon('/shutdown');
});
        </script>
    ''', data=data, person_data=person_data)

@app.route('/open-receipt-folder')
def open_receipt_folder():
    folder_path = os.path.join(os.getcwd(), 'receipt')
    os.startfile(folder_path)
    return redirect('/sales')

@app.route('/open-sale-folder')
def open_sale_folder():
    folder_path = os.path.dirname(os.path.abspath(sales_file_path))
    os.startfile(folder_path)
    return redirect('/sales')

receipt_folder = os.path.join(os.getcwd(), 'receipt')
os.makedirs(receipt_folder, exist_ok=True)

@app.route('/save_receipt', methods=['POST'])
def save_receipt():
    if 'receipt' not in request.files:
        return jsonify({'success': False, 'message': 'No receipt file part'})

    file = request.files['receipt']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'})

    now = datetime.now()
    formatted_date = now.strftime('%d%m%Y')
    formatted_time = now.strftime('%H%M%S')
    filename = f'receipt-{formatted_date}-{formatted_time}.png'
    file_path = os.path.join(receipt_folder, filename)
    file.save(file_path)
    return jsonify({'success': True})


@app.route('/view_receipt/<filename>')
def view_receipt(filename):
    return send_from_directory(receipt_folder, filename)

@app.route('/exit', methods=['GET'])
def exit_app():
    print('Exit signal sent.')
    os.kill(os.getpid(), signal.SIGINT)
    return 'Exiting...'

if __name__ == '__main__':

    flask_thread = threading.Thread(target=run_flask)
    flask_thread.start()

    webview.create_window("Market App", "http://127.0.0.1:5000/", fullscreen=False)
    webview.start()